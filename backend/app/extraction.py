"""ST.AIRS — Document Text Extraction"""

import csv
import io
import re


def extract_text(file_bytes: bytes, filename: str, content_type: str) -> tuple:
    """Extract text from a document. Returns (text, extra_metadata)."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext == "pdf" or content_type == "application/pdf":
            return _extract_pdf(file_bytes)
        elif ext == "docx" or "wordprocessingml" in content_type:
            return _extract_docx(file_bytes)
        elif ext == "xlsx" or "spreadsheetml" in content_type:
            return _extract_xlsx(file_bytes)
        elif ext == "csv" or content_type == "text/csv":
            return _extract_csv(file_bytes)
        elif ext == "txt" or content_type.startswith("text/"):
            return _extract_txt(file_bytes)
        elif content_type.startswith("image/"):
            return "", {"note": "Image file — no text extraction"}
        else:
            return "", {"note": "Unsupported format for extraction"}
    except Exception as e:
        return "extraction_failed", {"extraction_error": str(e)}


def clean_extracted_text(raw_text):
    """Clean raw extracted text for better readability.

    - Collapses multiple spaces into one
    - Removes orphan single characters (isolated letters from broken columns)
    - Adds line breaks between detected paragraphs
    - Strips leading/trailing whitespace from lines
    """
    if not raw_text or raw_text == "extraction_failed":
        return raw_text

    # Split into lines for processing
    lines = raw_text.split("\n")
    cleaned_lines = []

    for line in lines:
        # Collapse multiple spaces/tabs into single space
        line = re.sub(r"[ \t]+", " ", line).strip()
        # Remove lines that are just single isolated characters (orphans)
        if len(line) == 1 and not line.isdigit():
            continue
        # Remove lines that are only a series of isolated single chars separated by spaces
        tokens = line.split()
        if len(tokens) > 1 and all(len(t) == 1 for t in tokens):
            continue
        if line:
            cleaned_lines.append(line)

    # Rejoin and detect paragraph boundaries:
    # Insert double newlines where a line ends with sentence-ending punctuation
    # or where there's already a blank gap
    result_lines = []
    for i, line in enumerate(cleaned_lines):
        result_lines.append(line)
        # Add paragraph break after lines ending with sentence punctuation
        if line and line[-1] in ".!?:؟" and i < len(cleaned_lines) - 1:
            next_line = cleaned_lines[i + 1]
            # Only add break if next line starts with uppercase or is clearly new paragraph
            if next_line and (next_line[0].isupper() or next_line[0].isdigit()):
                result_lines.append("")

    return "\n".join(result_lines).strip()


def assess_extraction_quality(text):
    """Assess the quality of extracted text.

    Returns 'good' if text appears clean, 'partial' if it seems garbled.
    Heuristic: ratio of recognizable words (3+ chars, mostly letters) to total tokens.
    """
    if not text or text == "extraction_failed":
        return "failed"

    tokens = text.split()
    if len(tokens) < 5:
        return "good"  # Too short to judge meaningfully

    real_words = 0
    for token in tokens:
        # Strip common punctuation
        clean = re.sub(r"[^\w]", "", token)
        if len(clean) >= 2 and re.search(r"[a-zA-Z\u0600-\u06FF]", clean):
            real_words += 1

    ratio = real_words / len(tokens) if tokens else 0
    return "good" if ratio >= 0.5 else "partial"


def _extract_pdf(file_bytes: bytes) -> tuple:
    import pdfplumber
    pdf = pdfplumber.open(io.BytesIO(file_bytes))
    pages_text = []
    for page in pdf.pages:
        text = page.extract_text() or ""
        pages_text.append(text)
    pdf.close()

    raw_text = "\n\n".join(pages_text).strip()
    cleaned_text = clean_extracted_text(raw_text)
    quality = assess_extraction_quality(cleaned_text)

    return raw_text or "extraction_failed", {
        "page_count": len(pages_text),
        "pages_text": pages_text,
        "cleaned_text": cleaned_text,
        "extraction_quality": quality,
    }


def _extract_docx(file_bytes: bytes) -> tuple:
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    full_text = "\n".join(paragraphs).strip()
    return full_text or "extraction_failed", {"paragraph_count": len(paragraphs)}


def _extract_xlsx(file_bytes: bytes) -> tuple:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    sheet_count = len(wb.sheetnames)
    for sheet in wb.worksheets:
        lines.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
    wb.close()
    full_text = "\n".join(lines).strip()
    return full_text or "extraction_failed", {"sheet_count": sheet_count}


def _extract_csv(file_bytes: bytes) -> tuple:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    lines = ["\t".join(row) for row in rows]
    full_text = "\n".join(lines).strip()
    return full_text or "extraction_failed", {"row_count": len(rows)}


def _extract_txt(file_bytes: bytes) -> tuple:
    text = file_bytes.decode("utf-8", errors="replace").strip()
    return text or "extraction_failed", {}
